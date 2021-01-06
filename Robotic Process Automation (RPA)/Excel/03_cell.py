from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active          # Load the acrtiving sheet form the work book
ws.title = "RPASheet"

ws["A1"] = 1            # Put the value into specific cells
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6


print(ws["A1"])             # Printing out the information about A1 cell
print(ws["A1"].value)       # Printing out the value that in the A1 cell
print(ws["A10"].value)      # Print "None" if the cell does not have a value



# column = A(1), B(2), C(3), D(4), E(5), F(6), ...
# row = 1, 2, 3, 4, 5, 6, ...
print(ws.cell(column=1, row=1).value)   # Same as "A1"
print(ws.cell(column=2, row=1).value)   # Same as "B1"


c = ws.cell(column=3, row=1, value=10)  # column=3, row=1 --> "C1"
                                        # Put the value(10) int the C1 cell
print(c.value)                          # Printing the value in the C1 cell


# for x in range(1, 11):                                  # 10 Row
#     for y in range(1,11):                               # 10 Column
#         ws.cell(row=x, column=y, value=randint(0,100))  # Put the random value between 0~100


count = 1
for x in range(1, 11):                                  # 10 Row
    for y in range(1,11):                               # 10 Column
        ws.cell(row=x, column=y, value=count)           # Put the value in order
        count += 1

        

wb.save("sample.xlsx")                   # Save the file with name
wb.close()                               # Close the work book