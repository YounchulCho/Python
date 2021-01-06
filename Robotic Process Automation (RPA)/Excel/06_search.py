from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")       # sample.xlsx file load
ws = wb.active


for row in ws.iter_rows(min_row=2):     # ID, Eng, Math
    if int(row[1].value) > 80:
        print(row[0].value, "is genious of English")


# Change the value "Eng" -> "Com"
for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "Eng":
            cell.value = "Com"



wb.save("sample_modified.xlsx")         # Save the file with name
wb.close()                              # Close the work book