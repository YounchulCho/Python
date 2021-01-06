from openpyxl import load_workbook

wb = load_workbook("sample_merge.xlsx")     # sample.xlsx file load
ws = wb.active

ws.unmerge_cells("B2:D2")                   # Unmerge the cell

wb.save("sample_unmerge.xlsx")
wb.close()                                  # Close the work book