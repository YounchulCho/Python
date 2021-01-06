from openpyxl import Workbook

wb = wb = Workbook()                # Create the work book
ws = wb.active


ws.merge_cells("B2:D2")             # Merge the cells from B2 to D2
ws["B2"].value = "Merged Cell"      # Put the value into the merged cell



wb.save("sample_merge.xlsx")
wb.close()                           # Close the work book