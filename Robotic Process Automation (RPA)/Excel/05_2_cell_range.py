from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string

wb = Workbook()
ws = wb.active

ws.append(["ID", "Eng", "Math"])        # Put the values on the line
for i in range(1, 11):                  # Put 10 data
    ws.append([i, randint(0,100), randint(0,100)])




# All rows
# print(tuple(ws.rows))                   # It tuple by one line from the row
                                        #   (A1, B1, C1), (A2, B2, C2), (.....), ...
# The one way
for row in tuple(ws.rows):
    print(row)
    print(row[1].value)

# The the other way
for row in ws.iter_rows():
    print(row)
    print(row[1].value)


# All columns
# print(tuple(ws.columns))                # It tuple by one line from the columns
                                        #   (A1, A2, A3), (B1, B2, B3), (.....), ...
# The one way
for col in tuple(ws.columns):
    print(col)
    print(col[0].value)
# The the other way
for col in ws.iter_cols():
    print(col)
    print(col[0].value)




# Select the sepcific range
# Row
#   min_row, min_col, max_row, max_col
for row in ws.iter_rows(min_row=1, max_row=5):
    print(row[1].value)

# From 2nd to 1th Row, and 2nd to 3rd Column
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
    # print(row)
    print(row[0].value, row[1].value)     # Eng, Math


# Column
for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    # print(col)
    print(col[0].value)





wb.save("sample.xlsx")                  # Save the file with name
wb.close()                              # Close the work book