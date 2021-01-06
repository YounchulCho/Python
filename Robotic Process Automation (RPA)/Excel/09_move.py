from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")       # sample.xlsx file load
ws = wb.active


ws.move_range("B1:C11", rows=0, cols=1) # Move one step to the left the B and C columns
ws["B1"].value = "Com"                  # Add one more subject



ws.move_range("C1:C11", rows=5, cols=-1)    # C1-C11 data move at B6-B16



wb.save("sample_move.xlsx")
wb.close()                              # Close the work book