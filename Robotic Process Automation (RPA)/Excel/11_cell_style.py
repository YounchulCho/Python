from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

wb = load_workbook("sample.xlsx")       # sample.xlsx file load
ws = wb.active

# Set a style for the title
a1 = ws["A1"]           # ID
b1 = ws["B1"]           # Eng
c1 = ws["C1"]           # Math


ws.column_dimensions["A"].width = 5     # Change the width of cell
ws.row_dimensions[1].height = 50        # Change the height of cell



#### Change the font style (Change the font color, style, size, etc.)
a1.font = Font(color="ff0000", italic=True, bold=True)      # color:red, itialic, and bold
b1.font = Font(color="cc33ff", name="Arial", strike=True)   # color:purple, font:Arial, and strike
c1.font = Font(color="0000ff", size=20, underline='single') # color:purple, size:20, and single underline



#### Adding the Border line
# Set the border line style
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
a1.border = thin_border     # Apply the border style to cell(A1)
b1.border = thin_border     # Apply the border style to cell(A2)
c1.border = thin_border     # Apply the border style to cell(A3)



#### Change the Cell color
for row in ws.rows:
    for cell in row:
        if cell.column == 1:    # Skip the ID
            continue

        # The value in the cell is integer type and bigger than 90
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor="00ff00", fill_type="solid")    # Change background color
            cell.font = Font(color="ff0000")    # Change the font color



#### Chage the alignment
for row in ws.rows:
    for cell in row:
        # Center alignment for each cell
        cell.alignment = Alignment(horizontal="center", vertical="center")  # center, left, right, top, bottom




#### Freeze Panes
ws.freeze_panes = "B2"



wb.save("sample_style.xlsx")
wb.close()                              # Close the work book