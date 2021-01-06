from openpyxl import load_workbook

wb = load_workbook("sample_formula.xlsx")       # sample.xlsx file load
ws = wb.active

# Get the row data (forumla)
for row in ws.values:
    for cell in row:
        print(cell)



#### Load the file without formula
wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# It displays "None", if the data is not evaluated
# It shows the value, after open, save, and close the file (not in the code)
for row in ws.values:
    for cell in row:
        print(cell)



wb.close()                              # Close the work book