from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet()                      # Create new sheet
ws.title = "MySheet"                        # Change the new sheet's name
ws.sheet_properties.tabColor = "ff66ff"     # Change the tab color, it uses RGB color code

ws1 = wb.create_sheet("YourSheet")          # Create the new sheet with a name
ws2 = wb.create_sheet("NewSheet", 2)        # Create the new sheet and put the third postion
                                            #  -> The number is start from 0
new_ws = wb["NewSheet"]                     # Access the sheet like Dictionary type


print(wb.sheetnames)                        # Print out all sheet's name


new_ws["A1"] = "Test"                       # Put the value in a specific position(A1)
target = wb.copy_worksheet(new_ws)          # Copy the sheet
target.title = "Copied Sheet"               # Naming the copied sheet


wb.save("sample.xlsx")                      # Save the file with name
wb.close()                                  # Close the work book