import datetime
from openpyxl import Workbook

wb = wb = Workbook()             # Create the work book
ws = wb.active


# Put the value in the cell
ws["A1"] = datetime.datetime.today()    # today' date
ws["A2"] = "=SUM(1,2,3)"                # 1+2+3 = 6
ws["A3"] = "=AVERAGE(1,2,3)"            # (1+2+3) / 3 = 2


ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)"



wb.save("sample_formula.xlsx")
wb.close()                              # Close the work book