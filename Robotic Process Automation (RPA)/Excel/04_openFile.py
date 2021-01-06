from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")       # sample.xlsx file load
ws = wb.active



# Print all value in the sheet
#   max_row: get the maximun size of row
#   max_column: get the maximum size of column
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end = " ")
    print()



wb.close()                               # Close the work book