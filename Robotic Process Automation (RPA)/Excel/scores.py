####################
# 1. Put the final grade
#    Id, Attantion, Quize1, Quize2, Midterm, Final, Project
#    1,10,8,5,14,26,12
#    2,7,3,7,15,24,18
#    3,9,5,8,8,12,4
#    4,7,8,7,17,21,18
#    5,7,8,7,16,25,15
#    6,3,5,8,8,17,0
#    7,4,9,10,16,27,18
#    8,6,6,6,15,19,17
#    9,10,10,9,19,30,19
#    10,9,8,8,20,25,20
# 2. Found some error in Quize 2. Give 10 point to everyone
# 3. Add sum value(Total score) in "H", Grade information in "I"
#       - A: >= 90, B: >= 80, C: >= 70, else D
#   3.1. If attend is less than 5, grade is F
####################
from openpyxl import Workbook

wb = wb = Workbook()                # Create the work book
ws = wb.active

#####   Step 1
# Put the title information
ws.append(("Id", "Attend", "Quize1", "Quize2", "Midterm", "Final", "Project"))

scores = [
    (1,10,8,5,14,26,12),
    (2,7,3,7,15,24,18),
    (3,9,5,8,8,12,4),
    (4,7,8,7,17,21,18),
    (5,7,8,7,16,25,15),
    (6,3,5,8,8,17,0),
    (7,4,9,10,16,27,18),
    (8,6,6,6,15,19,17),
    (9,10,10,9,19,30,19),
    (10,9,8,8,20,25,20)
    ]

# Put every score value
for s in scores:
    ws.append(s)


#####   Step 2
for index, cell in enumerate(ws["D"]):
    if index == 0:      # Skip the title
        continue
    cell.value = 10     # Change the value to 10



#####   Step 3
ws["H1"] = "Total score"
ws["I1"] = "Grade"

for index, score in enumerate(scores, start=2): # It reads the data from 2 --> Skip the title
    # "H" is column 8
    ws.cell(row=index, column=8).value = "=SUM(B{}:G{})".format(index, index)


    # Sum all data -> subtract original Quiz 2 score -> plus 10
    # The data is not evaluated. It needs calculate itself
    sum_val = sum(score[1:]) - score[3] + 10
    grade = None
    if sum_val >= 90:
        grade = "A"
    elif sum_val >= 80:
        grade = "B"
    elif sum_val >= 70:
        grade = "C"
    else:
        grade = "D"    


##### Step 3.1
    if score[1] < 5:
        grade = "F"


    ws.cell(row=index, column=9).value = grade



wb.save("scores.xlsx")
wb.close()                           # Close the work book