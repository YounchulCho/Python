from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string


wb = Workbook()
ws = wb.active


ws.append(["ID", "Eng", "Math"])        # Put the values on the line
for i in range(1, 11):                  # Put 10 data
    ws.append([i, randint(0,100), randint(0,100)])


# Get by Column
col_B = ws["B"]                         # Get "B" column - Eng column
for cell in col_B:
    print(cell.value)


col_range = ws["B:C"]                   # Get information from "B" to "C" column
for cols in col_range:
    for cell in cols:
        print(cell.value)
    print()



# Get by Row
row_title = ws[1]                       # Get "1" row - title row
for cell in row_title:
    print(cell.value)


row_range = ws[2:6]                     # Get all grade information
for rows in row_range:
    for cell in rows:
        print(cell.value, end = " ")
    print()


row_range = ws[2:ws.max_row]            # Get from row 2 to end point
for rows in row_range:
    for cell in rows:
        print(cell.coordinate, end = " ")   # Print the cell information
        print(cell.value, end = " ")        # Print the value
    print()


row_range = ws[2:ws.max_row]            # Get from row 2 to end point
for rows in row_range:
    for cell in rows:
        xy = coordinate_from_string(cell.coordinate)    # Like Tuple, 
        print(xy, end = " ")                          # it has both information for cell and value
        # print(xy[0], end="")                            # A, B, C, D, ...
        # print(xy[1], end=" ")                           # 1, 2, 3, 4, ...
    print()



wb.save("sample.xlsx")                  # Save the file with name
wb.close()                              # Close the work book